import requests
from openpyxl import Workbook
from datetime import datetime


def get_random_users(count=5):
    url = f"https://randomuser.me/api/?results={count}"
    response = requests.get(url)
    
    if response.status_code == 200:
        data = response.json()
        users = data['results']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи"
        
        headers = ['№', 'Имя', 'Пол', 'Страна', 'Email', 'Телефон']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for i, user in enumerate(users, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=f"{user['name']['first']} {user['name']['last']}")
            ws.cell(row=row, column=3, value=user['gender'])
            ws.cell(row=row, column=4, value=user['location']['country'])
            ws.cell(row=row, column=5, value=user['email'])
            ws.cell(row=row, column=6, value=user['phone'])
        
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"random_users.xlsx"
        wb.save(filename)
        
        print("\nИнформация о случайных пользователях:")
        print("-" * 50)
        
        for i, user in enumerate(users, 1):
            print(f"\nПользователь #{i}:")
            print(f"Имя: {user['name']['first']} {user['name']['last']}")
            print(f"Пол: {user['gender']}")
            print(f"Страна: {user['location']['country']}")
            print(f"Email: {user['email']}")
            print(f"Телефон: {user['phone']}")
            print("-" * 50)
        
        print(f"\nДанные сохранены в файл: {filename}")
        
    else:
        print(f"Ошибка при получении данных: {response.status_code}")

if __name__ == "__main__":
    get_random_users(5)

